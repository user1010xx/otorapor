"""Performans raporu Excel üretimi."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from config import EXCEL_HEADERS


def build_performance_workbook(rows: list[dict[str, Any]]) -> BytesIO:
    """
    Görseldeki sütun düzeniyle .xlsx üretir.
    Süre alanları şüpheliyse ikinci sekmeye ham API alanlarını yazar (teşhis).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Performans"

    header_font = Font(bold=True)
    for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
            value = row.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if header in ("Dahili", "Dış Arama Sayısı") and isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
            elif header in ("Dış Arama Süresi", "Dış Arama Çaldırma Süresi"):
                cell.alignment = Alignment(horizontal="right")

    widths = {
        "A": 16,
        "B": 10,
        "C": 18,
        "D": 18,
        "E": 26,
    }
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    for col_idx in range(1, len(EXCEL_HEADERS) + 1):
        letter = get_column_letter(col_idx)
        if letter not in widths:
            ws.column_dimensions[letter].width = 14

    # Teşhis: ilk satırın ham alanları (süreler hâlâ şüpheliyse)
    sample = next((r for r in rows if r.get("_raw_sample")), None)
    if sample and _durations_look_suspicious(rows):
        dbg = wb.create_sheet("API_Ham_Alanlar")
        dbg.cell(1, 1, "alan").font = header_font
        dbg.cell(1, 2, "deger").font = header_font
        raw = sample.get("_raw_sample") or {}
        for i, (k, v) in enumerate(raw.items(), start=2):
            dbg.cell(i, 1, str(k))
            dbg.cell(i, 2, repr(v)[:500])
        dbg.column_dimensions["A"].width = 36
        dbg.column_dimensions["B"].width = 60

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _durations_look_suspicious(rows: list[dict[str, Any]]) -> bool:
    """Yüksek arama + 00:00:0x süre pattern'i."""
    suspicious = 0
    checked = 0
    for row in rows:
        try:
            count = int(row.get("Dış Arama Sayısı") or 0)
        except (TypeError, ValueError):
            count = 0
        talk = str(row.get("Dış Arama Süresi") or "")
        ring = str(row.get("Dış Arama Çaldırma Süresi") or "")
        if count < 10:
            continue
        checked += 1
        if talk.startswith("00:00:0") or ring.startswith("00:00:0"):
            suspicious += 1
    return checked > 0 and suspicious >= max(1, checked // 2)


def report_filename(when: datetime) -> str:
    return f"performans_raporu_{when.strftime('%Y-%m-%d_%H%M')}.xlsx"
