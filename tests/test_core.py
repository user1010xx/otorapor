"""Birim testleri — ağ / Telegram olmadan."""

from __future__ import annotations

from datetime import date, datetime, time
from io import BytesIO
from zoneinfo import ZoneInfo

import pytest
from openpyxl import load_workbook

from config import EXCEL_HEADERS, SCHEDULE_TIMES
from date_utils import DateParseError, join_command_args, parse_report_date
from excel_builder import build_performance_workbook, report_filename
from toniva_client import (
    _extract_rows,
    _format_duration,
    format_toniva_duration,
    normalize_row,
)


class TestDateParse:
    def test_dot_format(self):
        assert parse_report_date("18.07.2026") == date(2026, 7, 18)
        assert parse_report_date("1.7.2026") == date(2026, 7, 1)

    def test_short_year(self):
        assert parse_report_date("18.07.26") == date(2026, 7, 18)

    def test_iso(self):
        assert parse_report_date("2026-07-18") == date(2026, 7, 18)

    def test_slash(self):
        assert parse_report_date("18/07/2026") == date(2026, 7, 18)

    def test_invalid(self):
        with pytest.raises(DateParseError):
            parse_report_date("bugun")
        with pytest.raises(DateParseError):
            parse_report_date("32.13.2026")

    def test_join_args(self):
        assert join_command_args(["18.07.2026"]) == "18.07.2026"
        assert join_command_args([]) == ""
        assert join_command_args(None) == ""


class TestSchedule:
    def test_five_daily_slots(self):
        assert SCHEDULE_TIMES == [
            (12, 27),
            (13, 51),
            (16, 27),
            (17, 27),
            (18, 50),
        ]

    def test_times_are_valid_clock(self):
        for h, m in SCHEDULE_TIMES:
            t = time(hour=h, minute=m, tzinfo=ZoneInfo("Europe/Istanbul"))
            assert t.hour == h and t.minute == m


class TestDuration:
    def test_seconds_legacy(self):
        assert _format_duration(1908) == "00:31:48"
        assert _format_duration(4176) == "01:09:36"
        assert _format_duration(0) == "00:00:00"

    def test_hhmmss_text(self):
        assert _format_duration("1:09:36") == "01:09:36"
        assert _format_duration("00:39:36") == "00:39:36"
        assert _format_duration("39:36") == "00:39:36"

    def test_empty(self):
        assert _format_duration(None) == "00:00:00"
        assert _format_duration("") == "00:00:00"

    def test_toniva_hours_unit(self):
        """UI cl(e)=ti(floor(e*3600)) — süreler ondalık SAAT."""
        # 01:17:24 = 4644 sn = 1.29 saat
        assert format_toniva_duration(1.29, source_key="OutboundCallDuration") == "01:17:24"
        # 02:28:12 = 8892 sn = 2.47 saat
        assert format_toniva_duration(2.47, source_key="OutboundRingDuration") == "02:28:12"
        # 00:54:36 = 3276 sn = 0.91 saat
        assert format_toniva_duration(0.91, source_key="OutboundCallDuration") == "00:54:36"
        # 00:00:00
        assert format_toniva_duration(0, source_key="OutboundCallDuration") == "00:00:00"

    def test_was_bug_one_second_misread(self):
        """Eski bug: 1.29 saati 1 saniye sanıyordu → 00:00:01."""
        assert format_toniva_duration(1.29, source_key="OutboundCallDuration") != "00:00:01"
        assert format_toniva_duration(2.47, source_key="OutboundRingDuration") != "00:00:02"


class TestExtractRows:
    def test_list_payload(self):
        assert _extract_rows([{"a": 1}]) == [{"a": 1}]

    def test_rows_key(self):
        assert _extract_rows({"rows": [1, 2], "meta": {}}) == [1, 2]

    def test_data_capital(self):
        assert _extract_rows({"Status": True, "Data": [{"ExtensionName": "x"}]}) == [
            {"ExtensionName": "x"}
        ]

    def test_unknown(self):
        assert _extract_rows({"meta": {"ok": True}, "foo": "bar"}) == []


class TestNormalize:
    def test_toniva_real_fields_selcuk(self):
        """2. görsel (UI) selcuk satırı — ham saat birimi."""
        row = normalize_row(
            {
                "ExtensionName": "selcuk",
                "ExtensionNumber": 608,
                "OutboundCallCount": 891,
                "OutboundCallDuration": 1.29,  # 01:17:24
                "OutboundRingDuration": 2.47,  # 02:28:12
            }
        )
        assert row["Dahili Adı"] == "selcuk"
        assert row["Dahili"] == 608
        assert row["Dış Arama Sayısı"] == 891
        assert row["Dış Arama Süresi"] == "01:17:24"
        assert row["Dış Arama Çaldırma Süresi"] == "02:28:12"

    def test_toniva_real_fields_celal(self):
        """UI: celal 692 / 00:54:36 / 02:16:48"""
        # 00:54:36 = 3276/3600 = 0.91
        # 02:16:48 = 8208/3600 = 2.28
        row = normalize_row(
            {
                "ExtensionName": "celal",
                "ExtensionNumber": 632,
                "OutboundCallCount": 692,
                "OutboundCallDuration": 0.91,
                "OutboundRingDuration": 2.28,
            }
        )
        assert row["Dahili Adı"] == "celal"
        assert row["Dış Arama Sayısı"] == 692
        assert row["Dış Arama Süresi"] == "00:54:36"
        assert row["Dış Arama Çaldırma Süresi"] == "02:16:48"

    def test_not_constant_wrong_duration(self):
        """Herkeste 01:40:00 olmamalı — farklı süreler korunur."""
        a = normalize_row(
            {
                "ExtensionName": "a",
                "ExtensionNumber": 1,
                "OutboundCallCount": 100,
                "OutboundCallDuration": 1.29,
                "OutboundRingDuration": 2.47,
            }
        )
        b = normalize_row(
            {
                "ExtensionName": "b",
                "ExtensionNumber": 2,
                "OutboundCallCount": 100,
                "OutboundCallDuration": 0.91,
                "OutboundRingDuration": 2.28,
            }
        )
        assert a["Dış Arama Süresi"] != b["Dış Arama Süresi"]
        assert a["Dış Arama Süresi"] == "01:17:24"
        assert b["Dış Arama Süresi"] == "00:54:36"

    def test_hhmmss_string_passthrough(self):
        row = normalize_row(
            {
                "ExtensionName": "adem",
                "ExtensionNumber": 583,
                "OutboundCallCount": 365,
                "OutboundCallDuration": "00:31:48",
                "OutboundRingDuration": "01:09:36",
            }
        )
        assert row["Dış Arama Süresi"] == "00:31:48"
        assert row["Dış Arama Çaldırma Süresi"] == "01:09:36"

    def test_missing_fields_defaults(self):
        row = normalize_row({})
        assert row["Dahili Adı"] == ""
        assert row["Dahili"] == 0
        assert row["Dış Arama Sayısı"] == 0
        assert row["Dış Arama Süresi"] == "00:00:00"


class TestExcel:
    def test_headers_and_values(self):
        rows = [
            {
                "Dahili Adı": "selcuk",
                "Dahili": 608,
                "Dış Arama Sayısı": 891,
                "Dış Arama Süresi": "01:17:24",
                "Dış Arama Çaldırma Süresi": "02:28:12",
            }
        ]
        buf = build_performance_workbook(rows)
        assert isinstance(buf, BytesIO)
        assert buf.getvalue()[:2] == b"PK"

        wb = load_workbook(buf)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, 6)]
        assert headers == EXCEL_HEADERS
        assert ws.cell(2, 1).value == "selcuk"
        assert ws.cell(2, 4).value == "01:17:24"
        assert ws.cell(2, 5).value == "02:28:12"

    def test_empty_rows_still_has_header(self):
        buf = build_performance_workbook([])
        wb = load_workbook(buf)
        ws = wb.active
        assert [ws.cell(1, c).value for c in range(1, 6)] == EXCEL_HEADERS
        assert ws.cell(2, 1).value is None

    def test_filename(self):
        when = datetime(2026, 7, 18, 12, 27)
        assert report_filename(when) == "performans_raporu_2026-07-18_1227.xlsx"
        assert (
            report_filename(when, report_date=date(2026, 7, 18), full_day=True)
            == "performans_raporu_2026-07-18_tum_gun.xlsx"
        )


class TestBotGuards:
    def test_allowed_group(self):
        from bot_app import is_allowed_group
        from config import Settings

        settings = Settings(
            telegram_bot_token="x",
            telegram_chat_id=-100123,
            toniva_api_key="tva_x",
            toniva_base_url="https://crm.toniva.net/api/public/v1",
        )
        assert is_allowed_group(-100123, settings) is True
        assert is_allowed_group(-100999, settings) is False
        assert is_allowed_group(None, settings) is False


class TestSettings:
    def test_missing_env(self, monkeypatch):
        from config import Settings

        for key in (
            "TELEGRAM_BOT_TOKEN",
            "TELEGRAM_CHAT_ID",
            "TONIVA_API_KEY",
            "TONIVA_BASE_URL",
        ):
            monkeypatch.delenv(key, raising=False)

        with pytest.raises(RuntimeError, match="TELEGRAM_BOT_TOKEN|TELEGRAM_CHAT_ID"):
            Settings.from_env()

    def test_from_env_ok(self, monkeypatch):
        from config import Settings

        monkeypatch.setenv("TELEGRAM_BOT_TOKEN", "token")
        monkeypatch.setenv("TELEGRAM_CHAT_ID", "-10055")
        monkeypatch.setenv("TONIVA_API_KEY", "tva_test")
        monkeypatch.delenv("TONIVA_BASE_URL", raising=False)

        s = Settings.from_env()
        assert s.telegram_chat_id == -10055
        assert s.toniva_base_url.endswith("/api/public/v1")
